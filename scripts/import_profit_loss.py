#!/usr/bin/env python3
"""
Import Odoo Profit & Loss XLSX snapshot into Supabase.

Usage:  python scripts/import_profit_loss.py
        python scripts/import_profit_loss.py path/to/profit_and_loss.xlsx
"""
import os, sys, json, urllib.request, urllib.error

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
ENV_PATH = os.path.join(PROJECT_DIR, '.env.local')

env_vars = {}
with open(ENV_PATH, 'r') as f:
    for line in f:
        line = line.strip()
        if not line or line.startswith('#'): continue
        eq = line.index('=')
        env_vars[line[:eq].strip()] = line[eq+1:].strip()

SUPABASE_URL = env_vars.get('NEXT_PUBLIC_SUPABASE_URL', '')
SERVICE_KEY = env_vars.get('SUPABASE_SERVICE_ROLE_KEY', '')
if not SUPABASE_URL or not SERVICE_KEY:
    print("Missing env vars in .env.local"); sys.exit(1)

def sb_request(method, path, data=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    headers = {'apikey': SERVICE_KEY, 'Authorization': f'Bearer {SERVICE_KEY}',
               'Content-Type': 'application/json', 'Prefer': 'return=minimal'}
    body = json.dumps(data).encode('utf-8') if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp: return resp.status, resp.read().decode('utf-8')
    except urllib.error.HTTPError as e: return e.code, e.read().decode('utf-8')

# Known subtotal/total line names
SUBTOTALS = {'Gross Profit', 'Operating Income (or Loss)'}
TOTALS = {'Net Profit'}
HEADERS = {'Revenue', 'Less Costs of Revenue', 'Less Operating Expenses',
           'Plus Other Income', 'Less Other Expenses'}

def parse_xlsx(filepath):
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Profit and Loss']

    # Year from cell C1
    year_cell = ws.cell(1, 3).value
    snapshot_year = int(year_cell) if isinstance(year_cell, (int, float)) else int(str(year_cell).strip()) if str(year_cell).strip().isdigit() else 2025

    print(f"Report year: {snapshot_year}")
    rows = []
    row_order = 0

    for r in range(1, ws.max_row + 1):
        code = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        balance = ws.cell(r, 3).value

        # Skip header row and year row
        if name == 'Account Name' or (not name and not code): continue
        if not name and not code: continue

        code = str(code).strip() if code else None
        name = str(name).strip() if name else ''
        balance = float(balance) if isinstance(balance, (int, float)) and balance is not None else 0.0

        if not name: continue

        # Determine level
        if code:
            level = 'account'
        elif name in TOTALS:
            level = 'total'
        elif name in SUBTOTALS:
            level = 'subtotal'
        elif name in HEADERS:
            level = 'header'
        else:
            level = 'header'

        row_order += 1
        rows.append({
            'snapshot_year': snapshot_year,
            'row_order': row_order,
            'level': level,
            'code': code,
            'name': name,
            'balance': round(balance, 2),
        })

    return snapshot_year, rows

def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(PROJECT_DIR, 'OdooCsvFiles', 'profit_and_loss.xlsx')
    if not os.path.exists(xlsx_path): print(f"Not found: {xlsx_path}"); sys.exit(1)

    print(f"Reading: {xlsx_path}")
    snapshot_year, rows = parse_xlsx(xlsx_path)
    print(f"Parsed {len(rows)} rows\n")

    for row in rows:
        indent = '    ' if row['level'] == 'account' else '  ' if row['level'] in ('subtotal','total') else ''
        code_str = f"[{row['code']}] " if row['code'] else ''
        print(f"{indent}{code_str}{row['name']}: {row['balance']:,.2f}")
    print()

    print(f"Clearing existing P&L for {snapshot_year}...")
    status, body = sb_request('DELETE', f'odoo_pl_snapshots?snapshot_year=eq.{snapshot_year}')
    if status >= 300: print(f"Delete failed ({status}): {body}"); sys.exit(1)

    print(f"Inserting {len(rows)} rows...")
    status, body = sb_request('POST', 'odoo_pl_snapshots', rows)
    if status >= 300: print(f"Insert failed ({status}): {body}"); sys.exit(1)

    net = next((r['balance'] for r in rows if r['name'] == 'Net Profit'), 0)
    print(f"P&L for {snapshot_year} imported! Net Profit: {net:,.2f}")

if __name__ == '__main__':
    main()
