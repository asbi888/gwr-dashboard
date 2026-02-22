#!/usr/bin/env python3
"""
Import Odoo Aged Receivable XLSX snapshot into Supabase.

Usage:  python scripts/import_aged_receivable.py
        python scripts/import_aged_receivable.py path/to/aged_receivable.xlsx
"""
import os, sys, json, urllib.request, urllib.error
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
ENV_PATH = os.path.join(PROJECT_DIR, '.env.local')

env_vars = {}
with open(ENV_PATH, 'r') as f:
    for line in f:
        line = line.strip()
        if not line or line.startswith('#'): continue
        eq = line.index('=')
        env_vars[line[:eq].strip()] = line[eq+1:].strip()

SUPABASE_URL = env_vars.get('NEXT_PUBLIC_SUPABASE_URL', '')
SERVICE_KEY = env_vars.get('SUPABASE_SERVICE_ROLE_KEY', '')
if not SUPABASE_URL or not SERVICE_KEY:
    print("Missing env vars in .env.local"); sys.exit(1)

def sb_request(method, path, data=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    headers = {'apikey': SERVICE_KEY, 'Authorization': f'Bearer {SERVICE_KEY}',
               'Content-Type': 'application/json', 'Prefer': 'return=minimal'}
    body = json.dumps(data).encode('utf-8') if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp: return resp.status, resp.read().decode('utf-8')
    except urllib.error.HTTPError as e: return e.code, e.read().decode('utf-8')

def parse_xlsx(filepath):
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Aged Receivable']

    # Row 1 col B has "As of DD/MM/YYYY"
    date_cell = ws.cell(1, 2).value
    if isinstance(date_cell, str) and 'As of' in date_cell:
        date_str = date_cell.replace('As of ', '').strip()
        snapshot_date = datetime.strptime(date_str, '%d/%m/%Y').strftime('%Y-%m-%d')
    else:
        snapshot_date = datetime.now().strftime('%Y-%m-%d')

    print(f"Snapshot date: {snapshot_date}")

    # Column layout: A=Partner, B=Invoice Date, C=At Date, D=1-30, E=31-60, F=61-90, G=91-120, H=Older, I=Total
    rows = []
    row_order = 0

    for r in range(4, ws.max_row + 1):  # Data starts at row 4
        partner = ws.cell(r, 1).value
        if not partner: continue
        partner = str(partner).strip()

        def num(col):
            v = ws.cell(r, col).value
            return round(float(v), 2) if isinstance(v, (int, float)) and v is not None else 0.0

        is_total = (partner == 'Aged Receivable')
        row_order += 1

        rows.append({
            'snapshot_date': snapshot_date,
            'row_order': row_order,
            'partner_name': partner,
            'is_total': is_total,
            'at_date': num(3),
            'bucket_1_30': num(4),
            'bucket_31_60': num(5),
            'bucket_61_90': num(6),
            'bucket_91_120': num(7),
            'older': num(8),
            'total': num(9),
        })

    return snapshot_date, rows

def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(PROJECT_DIR, 'OdooCsvFiles', 'aged_receivable.xlsx')
    if not os.path.exists(xlsx_path): print(f"Not found: {xlsx_path}"); sys.exit(1)

    print(f"Reading: {xlsx_path}")
    snapshot_date, rows = parse_xlsx(xlsx_path)
    print(f"Parsed {len(rows)} rows\n")

    for row in rows:
        prefix = ">> " if row['is_total'] else "   "
        print(f"{prefix}{row['partner_name']:<40} Total: {row['total']:>14,.2f}")
    print()

    print(f"Clearing existing snapshot for {snapshot_date}...")
    status, body = sb_request('DELETE', f'odoo_ar_snapshots?snapshot_date=eq.{snapshot_date}')
    if status >= 300: print(f"Delete failed ({status}): {body}"); sys.exit(1)

    print(f"Inserting {len(rows)} rows...")
    status, body = sb_request('POST', 'odoo_ar_snapshots', rows)
    if status >= 300: print(f"Insert failed ({status}): {body}"); sys.exit(1)

    total_row = next((r for r in rows if r['is_total']), None)
    if total_row:
        print(f"Aged Receivable as of {snapshot_date} imported! Total: {total_row['total']:,.2f}")
    else:
        print(f"Aged Receivable as of {snapshot_date} imported! {len(rows)} partners")

if __name__ == '__main__':
    main()
