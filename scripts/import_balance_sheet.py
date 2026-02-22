#!/usr/bin/env python3
"""
Import Odoo Balance Sheet XLSX snapshot into Supabase.

Usage:  python scripts/import_balance_sheet.py
        python scripts/import_balance_sheet.py path/to/balance_sheet.xlsx

Reads Supabase credentials from .env.local (no extra dependencies).
"""
import os
import sys
import json
import urllib.request
import urllib.error
from datetime import datetime

# ── Load .env.local ──
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
ENV_PATH = os.path.join(PROJECT_DIR, '.env.local')

env_vars = {}
with open(ENV_PATH, 'r') as f:
    for line in f:
        line = line.strip()
        if not line or line.startswith('#'):
            continue
        eq = line.index('=')
        env_vars[line[:eq].strip()] = line[eq+1:].strip()

SUPABASE_URL = env_vars.get('NEXT_PUBLIC_SUPABASE_URL', '')
SERVICE_KEY = env_vars.get('SUPABASE_SERVICE_ROLE_KEY', '')

if not SUPABASE_URL or not SERVICE_KEY:
    print("Missing NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY in .env.local")
    sys.exit(1)


def supabase_request(method, path, data=None):
    """Make a request to Supabase REST API."""
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    headers = {
        'apikey': SERVICE_KEY,
        'Authorization': f'Bearer {SERVICE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=minimal',
    }
    body = json.dumps(data).encode('utf-8') if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            return resp.status, resp.read().decode('utf-8')
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode('utf-8')


def parse_xlsx(filepath):
    """Parse Odoo Balance Sheet XLSX into structured rows."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Balance Sheet']

    # Row 1 has the "As of" date in column C
    date_cell = ws.cell(1, 3).value  # e.g. "As of 21/02/2026"
    if isinstance(date_cell, str) and 'As of' in date_cell:
        date_str = date_cell.replace('As of ', '').strip()
        snapshot_date = datetime.strptime(date_str, '%d/%m/%Y').strftime('%Y-%m-%d')
    else:
        snapshot_date = datetime.now().strftime('%Y-%m-%d')

    print(f"Snapshot date: {snapshot_date}")

    # Parse the data rows (starting from row 3 which is the header row)
    rows = []
    current_section = None  # ASSETS, LIABILITIES, EQUITY
    row_order = 0

    for r in range(1, ws.max_row + 1):
        code = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        balance = ws.cell(r, 3).value

        # Skip empty rows and header row
        if not name and not code:
            continue
        if name == 'Account Name':
            continue

        # Clean values
        code = str(code).strip() if code else None
        name = str(name).strip() if name else ''
        balance = float(balance) if balance is not None else 0.0

        # Skip the final "LIABILITIES + EQUITY" total row
        if name == 'LIABILITIES + EQUITY':
            continue

        # Determine level and section
        if name in ('ASSETS', 'LIABILITIES', 'EQUITY'):
            current_section = name
            level = 'section'
        elif name.startswith('Plus '):
            # "Plus Fixed Assets", "Plus Non-current Assets", etc.
            level = 'category'
            name = name.replace('Plus ', '')
        elif code and code not in ('Code',):
            # Has an account code = individual account line
            level = 'account'
        elif name in ('Current Assets', 'Bank and Cash Accounts', 'Receivables',
                       'Prepayments', 'Current Liabilities', 'Payables',
                       'Non-current Liabilities', 'Unallocated Earnings',
                       'Retained Earnings'):
            # Known subcategory names
            level = 'subcategory'
        elif name in ('Current Year Unallocated Earnings',
                       'Previous Years Unallocated Earnings',
                       'Current Year Retained Earnings',
                       'Previous Years Retained Earnings'):
            level = 'account'
        else:
            level = 'category'

        row_order += 1
        rows.append({
            'snapshot_date': snapshot_date,
            'row_order': row_order,
            'level': level,
            'parent_section': current_section,
            'code': code,
            'name': name,
            'balance': round(balance, 2),
        })

    return snapshot_date, rows


def main():
    # Determine XLSX path
    if len(sys.argv) > 1:
        xlsx_path = sys.argv[1]
    else:
        xlsx_path = os.path.join(PROJECT_DIR, 'OdooCsvFiles', 'balance_sheet.xlsx')

    if not os.path.exists(xlsx_path):
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    print(f"Reading: {xlsx_path}")
    snapshot_date, rows = parse_xlsx(xlsx_path)
    print(f"Parsed {len(rows)} rows")
    print()

    # Show preview
    for row in rows:
        indent = '  ' if row['level'] == 'category' else '    ' if row['level'] == 'subcategory' else '      ' if row['level'] == 'account' else ''
        code_str = f"[{row['code']}] " if row['code'] else ''
        print(f"{indent}{code_str}{row['name']}: {row['balance']:,.2f}")
    print()

    # Delete existing snapshot for this date
    print(f"Clearing existing snapshot for {snapshot_date}...")
    status, body = supabase_request(
        'DELETE',
        f'odoo_bs_snapshots?snapshot_date=eq.{snapshot_date}'
    )
    if status >= 300:
        print(f"Delete failed ({status}): {body}")
        sys.exit(1)
    print("Done")

    # Insert new rows
    print(f"Inserting {len(rows)} rows...")
    status, body = supabase_request('POST', 'odoo_bs_snapshots', rows)
    if status >= 300:
        print(f"Insert failed ({status}): {body}")
        sys.exit(1)

    print(f"Snapshot for {snapshot_date} imported successfully!")
    print(f"  Rows: {len(rows)}")

    # Summary
    assets = sum(r['balance'] for r in rows if r['parent_section'] == 'ASSETS' and r['level'] == 'section')
    liab = sum(r['balance'] for r in rows if r['parent_section'] == 'LIABILITIES' and r['level'] == 'section')
    equity = sum(r['balance'] for r in rows if r['parent_section'] == 'EQUITY' and r['level'] == 'section')
    print(f"  Assets:      {assets:>15,.2f}")
    print(f"  Liabilities: {liab:>15,.2f}")
    print(f"  Equity:      {equity:>15,.2f}")


if __name__ == '__main__':
    main()
