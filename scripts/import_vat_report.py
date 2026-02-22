#!/usr/bin/env python3
"""
Import Odoo VAT3 Tax Report XLSX snapshot into Supabase.

Usage:  python scripts/import_vat_report.py
        python scripts/import_vat_report.py path/to/vat3_tax_report.xlsx

Reads Supabase credentials from .env.local (no extra dependencies).
"""
import os
import sys
import json
import urllib.request
import urllib.error
import re

# ── Load .env.local ──
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
ENV_PATH = os.path.join(PROJECT_DIR, '.env.local')

env_vars = {}
with open(ENV_PATH, 'r') as f:
    for line in f:
        line = line.strip()
        if not line or line.startswith('#'):
            continue
        eq = line.index('=')
        env_vars[line[:eq].strip()] = line[eq+1:].strip()

SUPABASE_URL = env_vars.get('NEXT_PUBLIC_SUPABASE_URL', '')
SERVICE_KEY = env_vars.get('SUPABASE_SERVICE_ROLE_KEY', '')

if not SUPABASE_URL or not SERVICE_KEY:
    print("Missing NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY in .env.local")
    sys.exit(1)


def supabase_request(method, path, data=None):
    """Make a request to Supabase REST API."""
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    headers = {
        'apikey': SERVICE_KEY,
        'Authorization': f'Bearer {SERVICE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=minimal',
    }
    body = json.dumps(data).encode('utf-8') if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            return resp.status, resp.read().decode('utf-8')
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode('utf-8')


# Section mapping
SECTION_MAP = {
    'OUTPUT': 'OUTPUT',
    'INPUT - Imports and Purchases': 'INPUT',
    'VAT ACCOUNT': 'VAT_ACCOUNT',
}

# Line numbers that are totals
TOTAL_LINES = {'5.', '9.', '10.', '14.', '15.3.', '16.', '17.', '18.', '19.'}

# Line numbers that are section sub-headers
SUBHEADER_LINES = {'1.', '6.', '8.', '15.'}


def parse_xlsx(filepath):
    """Parse Odoo VAT3 Tax Report XLSX into structured rows."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['VAT3 Tax Report']

    # Row 1 col B has the year
    year_cell = ws.cell(1, 2).value
    if isinstance(year_cell, (int, float)):
        snapshot_year = int(year_cell)
    elif isinstance(year_cell, str) and year_cell.strip().isdigit():
        snapshot_year = int(year_cell.strip())
    else:
        snapshot_year = 2025
        print(f"Warning: Could not parse year from '{year_cell}', defaulting to {snapshot_year}")

    print(f"Report year: {snapshot_year}")

    rows = []
    current_section = None
    row_order = 0

    for r in range(1, ws.max_row + 1):
        col_a = ws.cell(r, 1).value
        col_b = ws.cell(r, 2).value
        col_c = ws.cell(r, 3).value
        col_d = ws.cell(r, 4).value

        # Clean string values
        label = str(col_a).strip() if col_a else ''

        # Skip empty rows, header row, year row
        if not label and not col_b:
            continue
        if label == '' and col_b == 'Percent':
            continue

        # Check if this is a section header
        if label in SECTION_MAP:
            current_section = SECTION_MAP[label]
            row_order += 1
            rows.append({
                'snapshot_year': snapshot_year,
                'row_order': row_order,
                'level': 'section',
                'parent_section': current_section,
                'line_number': None,
                'name': label,
                'percent_value': None,
                'net_value': None,
                'vat_value': None,
            })
            continue

        # Skip if no section context yet
        if not current_section:
            continue

        # Extract line number from label
        line_match = re.match(r'^(\d+(?:\.\d+)*\.?)\s*(.*)', label)
        if line_match:
            line_number = line_match.group(1)
            if not line_number.endswith('.'):
                line_number += '.'
            name = line_match.group(2).strip()
        else:
            line_number = None
            name = label

        # Determine level
        if line_number in TOTAL_LINES:
            level = 'total'
        elif line_number in SUBHEADER_LINES:
            level = 'item'
        elif line_number and '.' in line_number.rstrip('.'):
            level = 'subitem'
        elif line_number:
            level = 'item'
        else:
            level = 'item'

        # Parse numeric values
        percent_val = float(col_b) if isinstance(col_b, (int, float)) else None
        net_val = float(col_c) if isinstance(col_c, (int, float)) else None
        vat_val = float(col_d) if isinstance(col_d, (int, float)) else None

        row_order += 1
        rows.append({
            'snapshot_year': snapshot_year,
            'row_order': row_order,
            'level': level,
            'parent_section': current_section,
            'line_number': line_number,
            'name': name,
            'percent_value': round(percent_val, 4) if percent_val is not None else None,
            'net_value': round(net_val, 2) if net_val is not None else None,
            'vat_value': round(vat_val, 2) if vat_val is not None else None,
        })

    return snapshot_year, rows


def main():
    # Determine XLSX path
    if len(sys.argv) > 1:
        xlsx_path = sys.argv[1]
    else:
        xlsx_path = os.path.join(PROJECT_DIR, 'OdooCsvFiles', 'vat3_tax_report.xlsx')

    if not os.path.exists(xlsx_path):
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    print(f"Reading: {xlsx_path}")
    snapshot_year, rows = parse_xlsx(xlsx_path)
    print(f"Parsed {len(rows)} rows")
    print()

    # Show preview
    for row in rows:
        indent = ''
        if row['level'] == 'item':
            indent = '  '
        elif row['level'] == 'subitem':
            indent = '    '
        elif row['level'] == 'total':
            indent = '  '

        ln = f"[{row['line_number']}] " if row['line_number'] else ''
        net = f"Value: {row['net_value']:>14,.2f}" if row['net_value'] is not None else ''
        vat = f"  VAT: {row['vat_value']:>14,.2f}" if row['vat_value'] is not None else ''
        print(f"{indent}{ln}{row['name']}  {net}{vat}")
    print()

    # Delete existing snapshot for this year
    print(f"Clearing existing snapshot for year {snapshot_year}...")
    status, body = supabase_request(
        'DELETE',
        f'odoo_vat_snapshots?snapshot_year=eq.{snapshot_year}'
    )
    if status >= 300:
        print(f"Delete failed ({status}): {body}")
        sys.exit(1)
    print("Done")

    # Insert new rows
    print(f"Inserting {len(rows)} rows...")
    status, body = supabase_request('POST', 'odoo_vat_snapshots', rows)
    if status >= 300:
        print(f"Insert failed ({status}): {body}")
        sys.exit(1)

    print(f"VAT report for {snapshot_year} imported successfully!")
    print(f"  Rows: {len(rows)}")

    # Key totals
    for row in rows:
        if row['line_number'] == '5.':
            print(f"  Output Total — Value: {row['net_value']:,.2f}  VAT: {row['vat_value']:,.2f}")
        elif row['line_number'] == '9.':
            print(f"  Input Total  — Value: {row['net_value']:,.2f}  VAT: {row['vat_value']:,.2f}")
        elif row['line_number'] == '19.':
            print(f"  VAT Payable  — VAT: {row['vat_value']:,.2f}")


if __name__ == '__main__':
    main()
